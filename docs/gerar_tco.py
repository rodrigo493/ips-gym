"""
Gera TCO_IPS_100academias.xlsx — comparativo Edge vs Data Center para 100 academias / 3 anos.
Executar: python docs/gerar_tco.py
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import os

OUT = os.path.join(os.path.dirname(__file__), "TCO_IPS_100academias.xlsx")

wb = openpyxl.Workbook()

# ── Cores
AZUL_ESC  = "103468"
AZUL_MED  = "1A569A"
AZUL_CLA  = "D6E4F7"
VERDE_ESC = "1B5E20"
VERDE_CLA = "D4EDDA"
GOLD      = "F0B429"
GOLD_CLA  = "FFF8DC"
CINZA     = "F5F5F5"
BRANCO    = "FFFFFF"
VERMELHO  = "C62828"
VERM_CLA  = "FDECEA"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom():
    s = Side(style="medium", color=AZUL_ESC)
    return Border(bottom=s)

def money(ws, row, col, value, color_fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = 'R$ #,##0'
    c.alignment = Alignment(horizontal="right")
    c.font = font(size=11)
    c.border = border_thin()
    if color_fill:
        c.fill = fill(color_fill)
    return c

def header_cell(ws, row, col, text, bg=AZUL_ESC, fg="FFFFFF", size=11, bold=True, wrap=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=bold, size=size, color=fg)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border = border_thin()
    return c

def label_cell(ws, row, col, text, bold=False, bg=BRANCO, indent=0):
    c = ws.cell(row=row, column=col, value=(" " * indent * 2) + text)
    c.font = font(bold=bold, size=11)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border_thin()
    return c

# ════════════════════════════════════════
# ABA 1 — TCO COMPARATIVO
# ════════════════════════════════════════
ws = wb.active
ws.title = "TCO Comparativo"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 22
ws.row_dimensions[1].height = 36
ws.row_dimensions[2].height = 20

# Título
ws.merge_cells("A1:D1")
c = ws["A1"]
c.value = "TCO — IPS Platform: Modelo Edge vs. Data Center"
c.font = font(bold=True, size=15, color=BRANCO)
c.fill = fill(AZUL_ESC)
c.alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A2:D2")
c = ws["A2"]
c.value = "100 academias · 3 câmeras/academia · Horizonte: 36 meses · Live Equipamentos Ltda."
c.font = font(italic=True, size=10, color="555555")
c.fill = fill(CINZA)
c.alignment = Alignment(horizontal="center", vertical="center")

# Cabeçalho de colunas
r = 4
header_cell(ws, r, 1, "Item de Custo", bg=AZUL_MED)
header_cell(ws, r, 2, "Modelo A\nEdge Pesado (atual)", bg="1B3A6B", wrap=True)
header_cell(ws, r, 3, "Modelo B\nData Center Híbrido", bg=VERDE_ESC, wrap=True)
header_cell(ws, r, 4, "Diferença (B − A)", bg="555555", wrap=True)
ws.row_dimensions[r].height = 36

def section(row, title):
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value=f"  {title}")
    c.font = font(bold=True, size=11, color=BRANCO)
    c.fill = fill(AZUL_MED)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22

def row_data(row, label, val_a, val_b, bg_a=BRANCO, bg_b=BRANCO, indent=0, bold=False):
    label_cell(ws, row, 1, label, bold=bold, indent=indent)
    money(ws, row, 2, val_a, bg_a)
    money(ws, row, 3, val_b, bg_b)
    diff = val_b - val_a
    c = money(ws, row, 4, diff, VERM_CLA if diff > 0 else VERDE_CLA)
    c.font = font(bold=True, color=VERMELHO if diff > 0 else VERDE_ESC)

# ── SEÇÃO 1: HARDWARE POR ACADEMIA
section(5, "1. HARDWARE POR UNIDADE (por academia, 1 vez)")
row_data(6,  "Processador de borda",          5800,  1200, indent=1)
row_data(7,  "  Modelo A: Jetson Orin NX 16GB",      5800,  0, bg_a=CINZA, indent=2)
row_data(8,  "  Modelo B: Raspberry Pi 5 + HAILO-8", 0,     1200, bg_b=CINZA, indent=2)
row_data(9,  "Câmeras HD (3 unidades)",        1200,  1200, indent=1)
row_data(10, "Cabeamento e instalação",         800,   600, indent=1)
row_data(11, "Display/tablet do aluno",         800,   800, indent=1)
row_data(12, "SUBTOTAL HARDWARE / ACADEMIA",   8600,  3800, bg_a=AZUL_CLA, bg_b=VERDE_CLA, bold=True)
row_data(13, "TOTAL HARDWARE 100 ACADEMIAS", 860000,380000, bg_a=AZUL_CLA, bg_b=VERDE_CLA, bold=True)

# ── SEÇÃO 2: DATA CENTER (apenas modelo B)
section(15, "2. INFRAESTRUTURA DATA CENTER — Modelo B (investimento inicial)")
label_cell(ws, 16, 1, "  2 servidores CPU (análise de regras + storage)", indent=1)
money(ws, 16, 2, 0)
money(ws, 16, 3, 20000, VERDE_CLA)
money(ws, 16, 4, 20000, VERM_CLA).font = font(bold=True, color=VERMELHO)

label_cell(ws, 17, 1, "  UPS + rack + switches + cabos", indent=1)
money(ws, 17, 2, 0)
money(ws, 17, 3, 15000, VERDE_CLA)
money(ws, 17, 4, 15000, VERM_CLA).font = font(bold=True, color=VERMELHO)

label_cell(ws, 18, 1, "  SUBTOTAL DATA CENTER SETUP", bold=True)
money(ws, 18, 2, 0, AZUL_CLA)
money(ws, 18, 3, 35000, VERDE_CLA)
money(ws, 18, 4, 35000, VERM_CLA).font = font(bold=True, color=VERMELHO)

# ── SEÇÃO 3: CUSTOS MENSAIS
section(20, "3. CUSTOS MENSAIS RECORRENTES (por mês)")
row_data(21, "Colocation + link 100Mbps dedicado",     0,    5000, indent=1)
row_data(22, "VPS backend (atual)",                   800,    300, indent=1)
row_data(23, "Storage / backup (Supabase/self-hosted)",200,    500, indent=1)
row_data(24, "Suporte técnico (custo menor pois\n  atualização é central)", 15000, 8000, indent=1)
row_data(25, "TOTAL MENSAL",                         16000,  13800, bg_a=AZUL_CLA, bg_b=VERDE_CLA, bold=True)

# ── SEÇÃO 4: TCO 36 MESES
section(27, "4. TCO TOTAL — 36 MESES")
row_data(28, "Hardware (academias)",              860000, 380000, indent=1)
row_data(29, "Data center setup",                      0,  35000, indent=1)
row_data(30, "Custos mensais × 36 meses",        576000, 496800, indent=1)
row_data(31, "TCO TOTAL 36 MESES",            1436000, 911800,
         bg_a=AZUL_CLA, bg_b=VERDE_CLA, bold=True)
row_data(32, "TCO POR ACADEMIA / MÊS",             399,    253,
         bg_a=AZUL_CLA, bg_b=VERDE_CLA, bold=True)

ws.row_dimensions[31].height = 24
ws.row_dimensions[32].height = 24

# Economia
ws.merge_cells("A34:D34")
c = ws["A34"]
c.value = f"  ✅  ECONOMIA TOTAL COM DATA CENTER:  R$ 524.200  (36% menor em 3 anos)"
c.font = Font(name="Arial", bold=True, size=13, color=BRANCO)
c.fill = fill(VERDE_ESC)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[34].height = 28

# ── SEÇÃO 5: RECEITA vs CUSTO
section(36, "5. RECEITA vs CUSTO POR ACADEMIA / MÊS")
row_data(37, "Mensalidade SaaS cobrada da academia",  400,   400, indent=1)
row_data(38, "Custo por academia / mês",              399,   253, indent=1)
row_data(39, "MARGEM BRUTA POR ACADEMIA",               1,   147,
         bg_a=VERM_CLA, bg_b=VERDE_CLA, bold=True)
row_data(40, "MARGEM BRUTA 100 ACADEMIAS / MÊS",       100, 14700,
         bg_a=VERM_CLA, bg_b=VERDE_CLA, bold=True)
row_data(41, "MARGEM BRUTA ANUAL",                   1200, 176400,
         bg_a=VERM_CLA, bg_b=VERDE_CLA, bold=True)

ws.merge_cells("A43:D43")
c = ws["A43"]
c.value = "  💡  Com data center: margem sobe de R$1 para R$147/academia/mês (+14.600%)"
c.font = Font(name="Arial", bold=True, size=12, color="7B3F00")
c.fill = fill(GOLD_CLA)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[43].height = 24

# Nota rodapé
ws.merge_cells("A45:D45")
c = ws["A45"]
c.value = "* Modelo B (Data Center Híbrido): edge device extrai keypoints (133 pontos) localmente via HAILO-8 e envia JSON ao data center. Sem streaming de vídeo. Latência estimada: 50-150ms. Fallback offline para funções básicas."
c.font = font(italic=True, size=9, color="666666")
c.fill = fill(CINZA)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[45].height = 32


# ════════════════════════════════════════
# ABA 2 — HARDWARE DETALHADO
# ════════════════════════════════════════
ws2 = wb.create_sheet("Hardware Detalhado")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 16
ws2.column_dimensions["E"].width = 18

ws2.merge_cells("A1:E1")
c = ws2["A1"]
c.value = "Hardware Detalhado — Modelo A (Edge) vs Modelo B (Data Center Híbrido)"
c.font = font(bold=True, size=14, color=BRANCO)
c.fill = fill(AZUL_ESC)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32

# Modelo A
ws2.merge_cells("A3:E3")
c = ws2["A3"]
c.value = "MODELO A — EDGE PESADO (Hardware por academia)"
c.font = font(bold=True, size=12, color=BRANCO)
c.fill = fill("1B3A6B")
c.alignment = Alignment(horizontal="center", vertical="center")

header_cell(ws2, 4, 1, "Componente", bg=AZUL_MED)
header_cell(ws2, 4, 2, "Especificação", bg=AZUL_MED)
header_cell(ws2, 4, 3, "Qtde", bg=AZUL_MED)
header_cell(ws2, 4, 4, "Unit. (R$)", bg=AZUL_MED)
header_cell(ws2, 4, 5, "Total (R$)", bg=AZUL_MED)

items_a = [
    ("Processador de borda", "NVIDIA Jetson Orin NX 16GB", 1, 5800),
    ("Câmera HD", "Arducam 2.1mm fisheye, USB", 3, 400),
    ("Cabeamento + fixação", "USB ativo 5m, suporte articulado", 3, 150),
    ("Gabinete protetivo", "ABS industrial, IP54", 1, 350),
    ("Display aluno", "Tablet 10\" Android", 1, 800),
    ("Instalação técnica", "Mão de obra + configuração", 1, 800),
]
r = 5
total_a = 0
for nome, spec, qtd, unit in items_a:
    total = qtd * unit
    total_a += total
    ws2.cell(r, 1, nome).font = font(size=11); ws2.cell(r,1).fill=fill(BRANCO); ws2.cell(r,1).border=border_thin()
    ws2.cell(r, 2, spec).font = font(size=10, color="444444"); ws2.cell(r,2).fill=fill(BRANCO); ws2.cell(r,2).border=border_thin()
    ws2.cell(r, 3, qtd).alignment = Alignment(horizontal="center"); ws2.cell(r,3).fill=fill(BRANCO); ws2.cell(r,3).border=border_thin()
    c4 = ws2.cell(r, 4, unit); c4.number_format='R$ #,##0'; c4.alignment=Alignment(horizontal="right"); c4.fill=fill(BRANCO); c4.border=border_thin()
    c5 = ws2.cell(r, 5, total); c5.number_format='R$ #,##0'; c5.alignment=Alignment(horizontal="right"); c5.fill=fill(BRANCO); c5.border=border_thin()
    r += 1

c = ws2.cell(r, 1, "TOTAL POR ACADEMIA"); c.font=font(bold=True,size=12); c.fill=fill(AZUL_CLA); c.border=border_thin()
ws2.merge_cells(f"A{r}:D{r}")
c5 = ws2.cell(r, 5, total_a); c5.number_format='R$ #,##0'; c5.font=font(bold=True,size=12,color=AZUL_ESC); c5.fill=fill(AZUL_CLA); c5.alignment=Alignment(horizontal="right"); c5.border=border_thin()

# Modelo B
r += 2
ws2.merge_cells(f"A{r}:E{r}")
c = ws2.cell(r, 1, "MODELO B — DATA CENTER HÍBRIDO (Hardware por academia)")
c.font = font(bold=True, size=12, color=BRANCO)
c.fill = fill(VERDE_ESC)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[r].height = 24
r += 1

header_cell(ws2, r, 1, "Componente", bg=VERDE_ESC)
header_cell(ws2, r, 2, "Especificação", bg=VERDE_ESC)
header_cell(ws2, r, 3, "Qtde", bg=VERDE_ESC)
header_cell(ws2, r, 4, "Unit. (R$)", bg=VERDE_ESC)
header_cell(ws2, r, 5, "Total (R$)", bg=VERDE_ESC)
r += 1

items_b = [
    ("Edge device + AI chip", "Raspberry Pi 5 8GB + HAILO-8 HAT+", 1, 1200),
    ("Câmera HD", "Câmera USB HD 1080p wide angle", 3, 400),
    ("Cabeamento + fixação", "USB ativo 5m, suporte articulado", 3, 150),
    ("Gabinete protetivo", "ABS industrial, IP54", 1, 250),
    ("Display aluno", "Tablet 10\" Android", 1, 800),
    ("Instalação técnica", "Mão de obra + configuração", 1, 600),
]
total_b_unit = 0
for nome, spec, qtd, unit in items_b:
    total = qtd * unit
    total_b_unit += total
    ws2.cell(r, 1, nome).font = font(size=11); ws2.cell(r,1).fill=fill(BRANCO); ws2.cell(r,1).border=border_thin()
    ws2.cell(r, 2, spec).font = font(size=10, color="444444"); ws2.cell(r,2).fill=fill(BRANCO); ws2.cell(r,2).border=border_thin()
    ws2.cell(r, 3, qtd).alignment = Alignment(horizontal="center"); ws2.cell(r,3).fill=fill(BRANCO); ws2.cell(r,3).border=border_thin()
    c4 = ws2.cell(r, 4, unit); c4.number_format='R$ #,##0'; c4.alignment=Alignment(horizontal="right"); c4.fill=fill(BRANCO); c4.border=border_thin()
    c5 = ws2.cell(r, 5, total); c5.number_format='R$ #,##0'; c5.alignment=Alignment(horizontal="right"); c5.fill=fill(BRANCO); c5.border=border_thin()
    r += 1

c = ws2.cell(r, 1, "TOTAL POR ACADEMIA"); c.font=font(bold=True,size=12); c.fill=fill(VERDE_CLA); c.border=border_thin()
ws2.merge_cells(f"A{r}:D{r}")
c5 = ws2.cell(r, 5, total_b_unit); c5.number_format='R$ #,##0'; c5.font=font(bold=True,size=12,color=VERDE_ESC); c5.fill=fill(VERDE_CLA); c5.alignment=Alignment(horizontal="right"); c5.border=border_thin()

# Data center setup
r += 2
ws2.merge_cells(f"A{r}:E{r}")
c = ws2.cell(r, 1, "DATA CENTER — Infraestrutura Central (investimento único para 100 academias)")
c.font = font(bold=True, size=12, color=BRANCO)
c.fill = fill("8B0000")
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[r].height = 24
r += 1

header_cell(ws2, r, 1, "Componente", bg="8B0000")
header_cell(ws2, r, 2, "Especificação", bg="8B0000")
header_cell(ws2, r, 3, "Qtde", bg="8B0000")
header_cell(ws2, r, 4, "Unit. (R$)", bg="8B0000")
header_cell(ws2, r, 5, "Total (R$)", bg="8B0000")
r += 1

items_dc = [
    ("Servidor de análise (CPU)", "AMD EPYC / Ryzen 9 + 64GB RAM + NVMe 4TB", 2, 10000),
    ("UPS rack 3kVA", "Proteção contra queda de energia", 1, 4000),
    ("Switch gerenciável 24p", "Rede interna do rack", 1, 2500),
    ("Rack 12U + organiz. cabos", "Estrutura física do data center", 1, 3500),
    ("Patch panel + cabeamento interno", "Cat6A blindado", 1, 2000),
    ("Instalação e configuração", "Setup inicial, hardening, monitoramento", 1, 3000),
]
total_dc = 0
for nome, spec, qtd, unit in items_dc:
    total = qtd * unit
    total_dc += total
    ws2.cell(r, 1, nome).font = font(size=11); ws2.cell(r,1).fill=fill(BRANCO); ws2.cell(r,1).border=border_thin()
    ws2.cell(r, 2, spec).font = font(size=10, color="444444"); ws2.cell(r,2).fill=fill(BRANCO); ws2.cell(r,2).border=border_thin()
    ws2.cell(r, 3, qtd).alignment = Alignment(horizontal="center"); ws2.cell(r,3).fill=fill(BRANCO); ws2.cell(r,3).border=border_thin()
    c4 = ws2.cell(r, 4, unit); c4.number_format='R$ #,##0'; c4.alignment=Alignment(horizontal="right"); c4.fill=fill(BRANCO); c4.border=border_thin()
    c5 = ws2.cell(r, 5, total); c5.number_format='R$ #,##0'; c5.alignment=Alignment(horizontal="right"); c5.fill=fill(BRANCO); c5.border=border_thin()
    r += 1

c = ws2.cell(r, 1, "TOTAL DATA CENTER SETUP"); c.font=font(bold=True,size=12); c.fill=fill(VERM_CLA); c.border=border_thin()
ws2.merge_cells(f"A{r}:D{r}")
c5 = ws2.cell(r, 5, total_dc); c5.number_format='R$ #,##0'; c5.font=font(bold=True,size=12,color=VERMELHO); c5.fill=fill(VERM_CLA); c5.alignment=Alignment(horizontal="right"); c5.border=border_thin()


# ════════════════════════════════════════
# ABA 3 — PROJEÇÃO FINANCEIRA
# ════════════════════════════════════════
ws3 = wb.create_sheet("Projeção Financeira")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 28

for col in range(2, 15):
    ws3.column_dimensions[get_column_letter(col)].width = 14

ws3.merge_cells("A1:N1")
c = ws3["A1"]
c.value = "Projeção Financeira — Crescimento de 10 para 100 academias (36 meses)"
c.font = font(bold=True, size=14, color=BRANCO)
c.fill = fill(AZUL_ESC)
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 32

# Cabeçalhos dos meses
meses = ["Mês 1", "Mês 3", "Mês 6", "Mês 9", "Mês 12",
         "Mês 15", "Mês 18", "Mês 21", "Mês 24",
         "Mês 27", "Mês 30", "Mês 33", "Mês 36"]

header_cell(ws3, 3, 1, "Métrica", bg=AZUL_MED)
for i, m in enumerate(meses):
    header_cell(ws3, 3, i+2, m, bg=AZUL_MED)

# Academias por mês (ramp-up)
academias = [10, 15, 25, 35, 45, 55, 65, 72, 80, 86, 92, 97, 100]

rows_proj = [
    ("Academias ativas",        academias,                        CINZA,    False),
    ("Receita mensal (R$400)",  [a*400 for a in academias],       VERDE_CLA, True),
    ("Custo mensal (B) R$138",  [a*138 for a in academias],       VERM_CLA,  True),
    ("Margem bruta mensal",     [a*(400-138) for a in academias], GOLD_CLA,  True),
    ("Custo acum. hardware B",  [a*3800 for a in academias],      AZUL_CLA,  True),
    ("Receita acumulada",       [sum([academias[j]*400 for j in range(i+1)]) for i in range(len(academias))], VERDE_CLA, True),
]

for ri, (label, vals, bg, is_money) in enumerate(rows_proj):
    r = 4 + ri
    label_cell(ws3, r, 1, label, bold=(ri in [1,3,5]))
    for ci, v in enumerate(vals):
        c = ws3.cell(r, ci+2, v)
        c.fill = fill(bg)
        c.border = border_thin()
        if is_money:
            c.number_format = 'R$ #,##0'
            c.alignment = Alignment(horizontal="right")
        else:
            c.alignment = Alignment(horizontal="center")
        c.font = font(size=10)

ws3.merge_cells("A11:N11")
c = ws3["A11"]
c.value = "* Custo mensal modelo B: R$138/academia = (R$5.800 data center + R$8.000 suporte) ÷ 100 academias. Amortiza hardware em 36 meses."
c.font = font(italic=True, size=9, color="666666")
c.fill = fill(CINZA)
c.alignment = Alignment(wrap_text=True, horizontal="left")
ws3.row_dimensions[11].height = 28

wb.save(OUT)
print(f"\nSalvo em: {OUT}")
print(f"\nResumo:")
print(f"  TCO Modelo A (Edge):         R$ 1.436.000")
print(f"  TCO Modelo B (Data Center):  R$   911.800")
print(f"  Economia:                    R$   524.200  (36%)")
print(f"  Margem/academia/mês Modelo A:  R$      1")
print(f"  Margem/academia/mês Modelo B:  R$    147")
